from flask import Flask, request, render_template, send_file
import face_recognition
import pandas as pd
import os
import json
from datetime import datetime
from flask import send_from_directory
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np
import cv2



app = Flask(__name__)

# Paths for uploads and reports
UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'reports'
ENCODINGS_FILE = 'encodings\student_encodings_almost_4.json'

# Ensure upload and report folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)

# Load student encodings from JSON
# with open(ENCODINGS_FILE, 'r') as f:
#     known_encodings = json.load(f)
with open(ENCODINGS_FILE, 'r') as f:
    student_encodings = json.load(f)

known_encodings = [
    {"name": name.strip(), "encoding": np.array(encoding)}
    for name, encoding in student_encodings.items()
]


@app.route('/')
def home():
    return render_template('index.html')

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        photos = request.files.getlist('photos')  # Get multiple uploaded files
        all_present_students = set()

        # Process each photo
        for photo in photos:
            path = os.path.join(UPLOAD_FOLDER, photo.filename)
            photo.save(path)

            # Get present students from each photo
            attendance_list = generate_attendance_list(path)
            all_present_students.update(attendance_list["present"])  # Combine present students

        # Determine all students and absent students
        all_students = {enc['name'] for enc in known_encodings}
        absent_students = all_students - all_present_students

        # Sort the lists
        present_students = sorted(all_present_students)
        absent_students = sorted(absent_students)

        # Generate report and pass sorted attendance lists
        report_path = save_attendance_report({
            "present": present_students,
            "absent": absent_students
        })

        # Render the attendance lists on the webpage
        return render_template(
            'attendance.html',
            present_students=present_students,
            absent_students=absent_students,
            report_path=f"/reports/{os.path.basename(report_path)}"
        )

    return render_template('upload.html')


@app.route('/reports/<filename>')
def download_report(filename):
    """Serve the attendance report for download."""
    return send_from_directory(REPORT_FOLDER, filename, as_attachment=True)

@app.route('/capture', methods=['GET', 'POST'])
def capture():
    if request.method == 'POST':
        # Save the captured photo
        photo = request.files['photo']
        path = os.path.join(UPLOAD_FOLDER, photo.filename)
        photo.save(path)

        # Process the photo and generate attendance
        attendance_list = generate_attendance_list(path)
        all_students = {enc['name'] for enc in known_encodings}
        absent_students = all_students - set(attendance_list["present"])

        # Generate the report
        report_path = save_attendance_report({
            "present": attendance_list["present"],
            "absent": list(absent_students)
        })

        # Display results
        return render_template(
            'attendance.html',
            present_students=sorted(attendance_list["present"]),
            absent_students=sorted(list(absent_students)),
            report_path=f"/reports/{os.path.basename(report_path)}"
        )

    return render_template('capture.html')


def generate_attendance_list(photo_path):
    print(f"Processing photo: {photo_path}")
    image = face_recognition.load_image_file(photo_path)
    face_locations = face_recognition.face_locations(image)  # Avoid resizing for debugging
    face_encodings = face_recognition.face_encodings(image, face_locations)

    print(f"Detected {len(face_locations)} faces in the image.")

    present_students = []
    for encoding in face_encodings:
        matches = face_recognition.compare_faces(
            [enc['encoding'] for enc in known_encodings], encoding, tolerance=0.5)
        if True in matches:
            match_index = matches.index(True)
            present_students.append(known_encodings[match_index]['name'])
        else:
            print("Unmatched face detected.")

    all_students = {enc['name'] for enc in known_encodings}
    absent_students = list(all_students - set(present_students))
    print(f"Present: {present_students}, Absent: {absent_students}")
    
    return {"present": present_students, "absent": absent_students}


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def save_attendance_report(attendance_list):
    """Save the attendance list as an Excel file with styled cells and adjusted sizes."""
    report_path = os.path.join(REPORT_FOLDER, f"attendance_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    
    # Create a new workbook and access the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Add headers
    sheet.append(["Name", "Status"])

    # Define the fill colors for absent and present students
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    present_fill = PatternFill(start_color="87F747", end_color="87F747", fill_type="solid")  # Green

    # Add data rows with styling
    for name in attendance_list["present"]:
        row_idx = sheet.max_row + 1
        sheet.append([name, "Present"])
        # Apply green fill to the 'Present' status cell
        sheet.cell(row=row_idx, column=2).fill = present_fill

    for name in attendance_list["absent"]:
        row_idx = sheet.max_row + 1
        sheet.append([name, "Absent"])
        # Apply light red fill to the 'Absent' status cell
        sheet.cell(row=row_idx, column=2).fill = absent_fill

    # Adjust column widths and row heights
    sheet.column_dimensions['A'].width = 30  # Adjust width for 'Name' column
    sheet.column_dimensions['B'].width = 15  # Adjust width for 'Status' column
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        sheet.row_dimensions[row[0].row].height = 20  # Adjust height for all rows

    # Save the workbook
    workbook.save(report_path)
    return report_path





if __name__ == '__main__':
    app.run(host='0.0.0.0',debug=True,port='5200')
